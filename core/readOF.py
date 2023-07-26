"""Модуль отвечает за выгрузку обменной формы из файла project."""

import datetime
import logging
import os
import pickle
import sys

import pandas as pd
import pythoncom
import win32com.client as win32
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

import settings.readOF as config
import settings.types as types

def get_project(path):
    """Открывает файл проекта и возвращает объект проекта.

    На вход поступает абсолютный путь до файла проекта. На выход
    поступает объект приложения MS Project и объект файла project.
    """
    if not os.path.isabs(path):
        logging.warning('%s: Путь до файла проекта не абсолютный', get_project.__name__)
    logging.info('%s: Пытаемся открыть файл проекта', get_project.__name__)
    try:
        msp = win32.Dispatch("MSProject.Application", pythoncom.CoInitialize())
        _abs_path = os.path.abspath(path)
        msp.FileOpen(_abs_path)
        project = msp.ActiveProject
    except Exception:
        logging.error('%s: Файл проекта не смог открыться', get_project.__name__)
        raise Exception('Не получилось открыть файл проекта')
    logging.info('%s: Файл проекта успешно открылся', get_project.__name__)
    return project, msp


def _get_data_task(t):
    """"Получает значения task из нужных столбцов.

    На вход поступает объект Task из project. Функция проходит и записывает
    в список те ячейки Task, которые соответствуют столбцам, указанным в конфиге.
    На выход поступает список полученных значений.
    """
    arr = []
    try:
        for i in config.ID_COLUMN.keys():
            try:
                data = getattr(t, i)
            except Exception as e:
                arr.append("Ошибка чтения")
                logging.error(f"{_get_data_task.__name__}: Возможно битая ячейка в столбце: {config.ID_COLUMN.get(i)}, c УИД: {arr[0]}"
                              )
                continue
            if isinstance(data, datetime.datetime):
                data = datetime.datetime.date(data)
            arr.append(data)
    except Exception as e:
        print(e)
        logging.error('%s: Неверный идентификатор столбца project', _get_data_task.__name__)
        raise Exception('Неверный идентификатор столбца project')
    return arr


def _get_column_name(worksheet):
    """Получает названия столбцов обменной формы.

    Это требуется так как столбцы в excel, прочитанные с помощью openpyxl, представляют собой объекты.
    Для упрощения в дальнейшем написания кода, эта функция достает из Excel названия столбцов.
    """
    column_headers = []
    for cell in worksheet[1]:
        column_headers.append(cell.value)
    return column_headers


def _get_column_to_switch_format(ws, column_names):
    """Создает словарь с объектами столбцов Excel и их требуемым типом и возвращает его.

    Столбцы в excel, прочитанные с помощью openpyxl, представляют собой объекты.
    Так как в словаре в конфиге прописаны названия столбцов, а не сами объекты, то
    мне нужно соотнести названия с объектами. Данная функция проходит по списку названий столбцов,
    которые нужно изменить, затем по списку названий всех столбцов, чтобы найти индексы
    нужных и по ним получить объекты столбцов. На выход поступает словарь, в качестве ключей содержит
    объекты столбцов excel, в качестве значений требуемый тип.
    """
    column_to_switch = {}
    ws_columns_list = list(ws.columns)
    for key in types.COLUMN_TYPES.keys():
        for i in range(1, len(column_names)):
            if column_names[i] == key:
                column_to_switch[ws_columns_list[i]] = types.COLUMN_TYPES[key]
    return column_to_switch


def fill_dataframe(project):
    """Заполняет DataFrame значениями из project.

    На вход поступают объект файла project. Из него
    формируется dataframe с данными из столбцов, которые указаны
    в конфиге и dataframe возвращается для дальнейшего использования.
    """
    logging.info('%s: Создаем DataFrame из столбцов объекта проекта', fill_dataframe.__name__)
    if not project:
        logging.error('%s: Не удалось получить объект проекта', fill_dataframe.__name__)
        raise Exception("Объект проекта пустой")
    if not config.ID_COLUMN:
        logging.error('%s: Ключевые столбцы не заданы', fill_dataframe.__name__)
        raise Exception("Ключевые столбцы не заданы")
    task_collection = project.Tasks
    data = pd.DataFrame(columns=config.ID_COLUMN.values())
    try:
        for t in task_collection:
            str = _get_data_task(t)
            data.loc[len(data.index)] = str
            if "Ошибка чтения" in str:
                raise Exception("Битая ячейка")
    except Exception:
        logging.error('%s: Ошибка при создании датафрейма', fill_dataframe.__name__)
        raise Exception("Ошибка при создании датафрейма")
    logging.info('%s: DataFrame из столбцов объекта проекта успешно создан', fill_dataframe.__name__)

    return data

def _resource_path(relative):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative)
    else:
        return os.path.join(os.path.abspath("."), relative)

def set_style_excel(column_index, path_to_excel):
    """Применяет стили к строкам excel.

    На вход поступают индекс колонки в Excel, которая является
    ключевой для задания стиля всему task, и путь до Excel.
    Функция создает объект worksheet из объекта workbook и
    применяет изменения из pickle файла, в котором содержится
    словарь для стилей для каждого конкретного ключевого значения
    (Фаза, феха и т.д.). Также она применяет тип данных там, где
    это требуется, например Дата и Процент.
    """
    logging.info("%s: Начало внедрения стилей в Excel файл", set_style_excel.__name__)
    try:
        with open(_resource_path(config.PATH_TO_STYLE_FILE), 'rb') as file:
            styles_dict = pickle.load(file)
    except FileNotFoundError:
        logging.error('%s: Неверно задан путь к файлу со стилями', set_style_excel.__name__)
        raise Exception("Неверный путь до файла со стилями")
    workbook_other = openpyxl.load_workbook(path_to_excel)
    worksheet_other = workbook_other.active
    headers = _get_column_name(worksheet_other)
    column_to_switch = _get_column_to_switch_format(worksheet_other, headers)
    for row in worksheet_other.iter_rows(min_row=1):
        cell = row[column_index - 1]
        cell_value = cell.value
        if cell_value in styles_dict:
            for cell in row:
                style = styles_dict[cell_value]
                cell.style = style
                column_letter = get_column_letter(cell.column)
                text_length = len(str(cell.value))
                current_width = worksheet_other.column_dimensions[column_letter].width
                if text_length > current_width:
                    worksheet_other.column_dimensions[column_letter].width = text_length

    for column_index, column in enumerate(worksheet_other.columns, start=1):
        column_iter = iter(column)
        next(column_iter)
        for cell in column_iter:
            if column in column_to_switch.keys():
                if column_to_switch[column] == "%":
                    cell.number_format = "0.00%"
                    cell.value = cell.value/100
            if isinstance(cell.value, datetime.date):
                cell.number_format = numbers.builtin_format_code(14)

    logging.info("%s: Стили успешно применены", set_style_excel.__name__)
    workbook_other.save(path_to_excel)



def main(path_to_project, path_to_folder):
    """Управляющая функция для контроллера.

    На вход поступает путь до файла project и до файла Excel.
    Выполняется выгрузка обменной формы. В качестве результата
    возвращается абсолютный путь до ОФ.
    """
    path_to_excel = None
    logging.basicConfig(level=logging.INFO, filename=config.PATH_TO_RESERVE_FOLDER+"//лог.log", filemode="w",
                        format="%(asctime)s %(levelname)s %(message)s")
    logging.info(f"Начало выгрузки обменной формы для файла: {path_to_project}")
    try:
        project, msp = get_project(path_to_project)
        data = fill_dataframe(project)
        file_name = os.path.splitext(os.path.basename(path_to_project))[0]
        current_date = datetime.datetime.now().strftime("%d.%m.%Y")
        path_to_excel = path_to_folder + "//" + file_name + "_ОФ_" + current_date + ".xlsx"
        data.to_excel(path_to_excel, sheet_name=f"Обменная форма {datetime.date.today()}", index=False)
        column_index = data.columns.get_loc(config.ID_COLUMN['Text5']) + 1
        set_style_excel(column_index, path_to_excel)
    except Exception as e:
        logging.error(f"Ошибка в выгрузке обменной формы из файла: {path_to_project}")
        return None
    logging.info(f"Обменная форма успешно выгружена, файл: {path_to_project}")
    msp.Quit()
    return path_to_excel
